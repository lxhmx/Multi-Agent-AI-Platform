"""
员工考勤扣款管理 API
提供考勤扣款数据的上传、查询、统计功能
"""
import sys
import io
from datetime import datetime
from pathlib import Path
from typing import Optional, List

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 添加项目根目录到路径
PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from database.mysql_client import MySQLClient

# 创建路由器
router = APIRouter(prefix="/api/financial/attendance", tags=["考勤扣款管理"])


# ==================== 工具函数 ====================

def extract_month_from_filename(filename: str) -> Optional[str]:
    """从文件名中提取月份信息"""
    import re
    
    patterns = [
        r'(\d{4})年(\d{1,2})月',
        r'(\d{4})-(\d{1,2})',
        r'(\d{4})(\d{2})',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            year, month = match.groups()
            return f"{year}-{int(month):02d}"
    
    month_match = re.search(r'(\d{1,2})月', filename)
    if month_match:
        month = int(month_match.group(1))
        year = datetime.now().year
        return f"{year}-{month:02d}"
    
    return None


def extract_month_from_excel(excel_path: str) -> Optional[str]:
    """从Excel表头提取考勤月份"""
    import re
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # 遍历前3行，查找包含"考勤"或年月信息的单元格
        for row_idx in range(1, 4):
            for col_idx in range(1, min(20, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value)
                    # 匹配 "2025年12月考勤" 或 "2025年12月" 格式
                    match = re.search(r'(\d{4})年(\d{1,2})月', cell_str)
                    if match:
                        year, month = match.groups()
                        wb.close()
                        return f"{year}-{int(month):02d}"
        
        wb.close()
    except Exception as e:
        print(f"[Attendance] 从Excel提取月份失败: {e}")
    
    return None


def process_attendance_deduction(excel_path: str):
    """处理考勤Excel文件，计算扣款"""
    # 导入考勤计算器
    sys.path.insert(0, str(PROJECT_ROOT / 'scripts'))
    from attendance_calculator import AttendanceCalculator
    
    calculator = AttendanceCalculator(excel_path)
    calculator.calculate_all()
    
    return calculator.results


def save_deduction_records(records: List[dict], attendance_month: str) -> dict:
    """保存扣款记录到数据库"""
    if not records:
        return {"inserted": 0, "updated": 0}
    
    inserted = 0
    updated = 0
    
    upsert_sql = """
    INSERT INTO employee_attendance_deduction 
        (employee_name, job_title, job_level, level_type, total_late_count,
         late_within_10_count, late_over_10_count, late_over_60_count,
         morning_missing_count, evening_missing_count, early_leave_count,
         total_deduction, attendance_month, row_order)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ON DUPLICATE KEY UPDATE
        job_title = VALUES(job_title),
        job_level = VALUES(job_level),
        level_type = VALUES(level_type),
        total_late_count = VALUES(total_late_count),
        late_within_10_count = VALUES(late_within_10_count),
        late_over_10_count = VALUES(late_over_10_count),
        late_over_60_count = VALUES(late_over_60_count),
        morning_missing_count = VALUES(morning_missing_count),
        evening_missing_count = VALUES(evening_missing_count),
        early_leave_count = VALUES(early_leave_count),
        total_deduction = VALUES(total_deduction),
        row_order = VALUES(row_order),
        updated_at = CURRENT_TIMESTAMP
    """
    
    try:
        with MySQLClient() as db:
            for idx, record in enumerate(records):
                params = (
                    record['name'],
                    record['position'],
                    record.get('level', ''),
                    record['level_prefix'],
                    record['total_late_count'],
                    record['late_within_10_count'],
                    record['late_over_10_count'],
                    record['late_over_60_count'],
                    record['morning_missing_count'],
                    record['evening_missing_count'],
                    record['early_leave_count'],
                    record['total_deduction'],
                    attendance_month,
                    idx
                )
                affected = db.execute_update(upsert_sql, params)
                if affected == 1:
                    inserted += 1
                else:
                    updated += 1
        
        return {"inserted": inserted, "updated": updated}
    except Exception as e:
        raise Exception(f"保存数据失败: {str(e)}")


# ==================== API 接口 ====================

@router.post("/upload")
async def upload_attendance_deduction(
    file: UploadFile = File(...),
    month: Optional[str] = Form(None)
):
    """上传考勤表并计算扣款数据"""
    try:
        if not file.filename.endswith(('.xls', '.xlsx')):
            raise HTTPException(status_code=400, detail="请上传Excel文件(.xls或.xlsx)")
        
        temp_dir = PROJECT_ROOT / 'temp'
        temp_dir.mkdir(exist_ok=True)
        temp_path = temp_dir / f"attendance_deduction_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
        
        content = await file.read()
        with open(temp_path, 'wb') as f:
            f.write(content)
        
        try:
            # 优先从Excel表头提取月份，其次从文件名，最后使用用户指定或当前月份
            attendance_month = month
            if not attendance_month:
                attendance_month = extract_month_from_excel(str(temp_path))
            if not attendance_month:
                attendance_month = extract_month_from_filename(file.filename)
            if not attendance_month:
                attendance_month = datetime.now().strftime('%Y-%m')
            
            records = process_attendance_deduction(str(temp_path))
            
            if not records:
                raise HTTPException(status_code=400, detail="未能从文件中解析出有效数据")
            
            result = save_deduction_records(records, attendance_month)
            
            total_deduction = sum(r['total_deduction'] for r in records)
            deduction_count = len([r for r in records if r['total_deduction'] > 0])
            total_late = sum(r['total_late_count'] for r in records)
            
            return {
                "success": True,
                "message": "考勤扣款数据处理完成",
                "data": {
                    "month": attendance_month,
                    "total_employees": len(records),
                    "deduction_employees": deduction_count,
                    "total_deduction": total_deduction,
                    "total_late_count": total_late,
                    "inserted": result["inserted"],
                    "updated": result["updated"]
                }
            }
        finally:
            if temp_path.exists():
                temp_path.unlink()
                
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"[Attendance] 上传处理失败: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")


@router.get("/records")
async def get_deduction_records(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    month: Optional[str] = Query(None, description="考勤月份(YYYY-MM)"),
    keyword: Optional[str] = Query(None, description="搜索关键词(姓名/职务)"),
    sort_field: Optional[str] = Query(None, description="排序字段"),
    sort_order: Optional[str] = Query(None, description="排序方向(asc/desc)")
):
    """获取考勤扣款记录列表"""
    try:
        conditions = []
        params = []
        
        if month:
            conditions.append("attendance_month = %s")
            params.append(month)
        
        if keyword:
            conditions.append("(employee_name LIKE %s OR job_title LIKE %s)")
            params.extend([f"%{keyword}%", f"%{keyword}%"])
        
        where_clause = " AND ".join(conditions) if conditions else "1=1"
        
        # 构建排序子句
        order_clause = "attendance_month DESC, row_order ASC"
        if sort_field and sort_order:
            allowed_fields = ['total_deduction', 'total_late_count', 'morning_missing_count', 
                            'evening_missing_count', 'early_leave_count', 'employee_name']
            if sort_field in allowed_fields:
                order_direction = 'ASC' if sort_order.lower() == 'asc' else 'DESC'
                order_clause = f"{sort_field} {order_direction}, row_order ASC"
        
        count_sql = f"SELECT COUNT(*) as total FROM employee_attendance_deduction WHERE {where_clause}"
        
        offset = (page - 1) * page_size
        data_sql = f"""
            SELECT * FROM employee_attendance_deduction 
            WHERE {where_clause}
            ORDER BY {order_clause}
            LIMIT %s OFFSET %s
        """
        
        with MySQLClient() as db:
            count_result = db.execute_query(count_sql, tuple(params))
            total = count_result[0]['total'] if count_result else 0
            
            data_params = tuple(params) + (page_size, offset)
            records = db.execute_query(data_sql, data_params)
        
        return {
            "success": True,
            "data": records,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": (total + page_size - 1) // page_size
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


@router.get("/stats")
async def get_deduction_stats(
    month: Optional[str] = Query(None, description="考勤月份(YYYY-MM)")
):
    """获取考勤扣款统计数据"""
    try:
        with MySQLClient() as db:
            month_condition = "WHERE attendance_month = %s" if month else ""
            month_params = (month,) if month else None
            
            stats_sql = f"""
                SELECT 
                    COUNT(*) as total_employees,
                    SUM(CASE WHEN total_deduction > 0 THEN 1 ELSE 0 END) as deduction_employees,
                    SUM(total_late_count) as total_late_count,
                    SUM(late_within_10_count) as late_within_10_count,
                    SUM(late_over_10_count) as late_over_10_count,
                    SUM(late_over_60_count) as late_over_60_count,
                    SUM(morning_missing_count) as morning_missing_count,
                    SUM(evening_missing_count) as evening_missing_count,
                    SUM(early_leave_count) as early_leave_count,
                    SUM(total_deduction) as total_deduction
                FROM employee_attendance_deduction
                {month_condition}
            """
            stats = db.execute_query(stats_sql, month_params)
            
            months_sql = """
                SELECT DISTINCT attendance_month 
                FROM employee_attendance_deduction 
                ORDER BY attendance_month DESC
            """
            months = db.execute_query(months_sql)
            
            # 按职级类型统计
            level_stats_sql = f"""
                SELECT 
                    level_type,
                    COUNT(*) as count,
                    SUM(total_late_count) as total_late,
                    SUM(total_deduction) as total_deduction
                FROM employee_attendance_deduction
                {month_condition}
                GROUP BY level_type
                ORDER BY FIELD(level_type, 'P', 'M', 'D', '实习')
            """
            level_stats = db.execute_query(level_stats_sql, month_params)
            
            # 扣款TOP10
            top_sql = f"""
                SELECT employee_name, job_title, level_type, total_late_count, total_deduction
                FROM employee_attendance_deduction
                {month_condition}
                ORDER BY total_deduction DESC
                LIMIT 10
            """
            top_employees = db.execute_query(top_sql, month_params)
        
        return {
            "success": True,
            "data": {
                "stats": stats[0] if stats else {},
                "months": [m['attendance_month'] for m in months],
                "level_stats": level_stats,
                "top_employees": top_employees
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"统计失败: {str(e)}")


@router.delete("/records")
async def delete_deduction_records(
    ids: Optional[List[int]] = Query(None, description="要删除的记录ID列表"),
    month: Optional[str] = Query(None, description="要删除的月份数据")
):
    """删除考勤扣款记录"""
    try:
        if not ids and not month:
            raise HTTPException(status_code=400, detail="请提供要删除的记录ID或月份")
        
        with MySQLClient() as db:
            if ids:
                placeholders = ','.join(['%s'] * len(ids))
                sql = f"DELETE FROM employee_attendance_deduction WHERE id IN ({placeholders})"
                affected = db.execute_update(sql, tuple(ids))
            else:
                sql = "DELETE FROM employee_attendance_deduction WHERE attendance_month = %s"
                affected = db.execute_update(sql, (month,))
        
        return {
            "success": True,
            "message": f"成功删除 {affected} 条记录"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


@router.get("/export")
async def export_deduction_records(
    month: Optional[str] = Query(None, description="考勤月份(YYYY-MM)"),
    keyword: Optional[str] = Query(None, description="搜索关键词")
):
    """导出考勤扣款记录为Excel"""
    try:
        conditions = []
        params = []
        
        if month:
            conditions.append("attendance_month = %s")
            params.append(month)
        
        if keyword:
            conditions.append("(employee_name LIKE %s OR job_title LIKE %s)")
            params.extend([f"%{keyword}%", f"%{keyword}%"])
        
        where_clause = " AND ".join(conditions) if conditions else "1=1"
        
        data_sql = f"""
            SELECT * FROM employee_attendance_deduction 
            WHERE {where_clause}
            ORDER BY attendance_month DESC, row_order ASC
        """
        
        with MySQLClient() as db:
            records = db.execute_query(data_sql, tuple(params))
        
        if not records:
            raise HTTPException(status_code=404, detail="没有数据可导出")
        
        # 创建Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "考勤扣款记录"
        
        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 表头
        headers = ['序号', '姓名', '职务', '职级', '职级类型', '总迟到', '10分内', 
                   '超10分', '超1小时', '早缺卡', '晚缺卡', '早退', '扣款金额', '考勤月份']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 数据行
        for idx, record in enumerate(records, 1):
            ws.cell(row=idx+1, column=1, value=idx).border = thin_border
            ws.cell(row=idx+1, column=2, value=record.get('employee_name', '')).border = thin_border
            ws.cell(row=idx+1, column=3, value=record.get('job_title', '')).border = thin_border
            ws.cell(row=idx+1, column=4, value=record.get('job_level', '')).border = thin_border
            ws.cell(row=idx+1, column=5, value=record.get('level_type', '')).border = thin_border
            ws.cell(row=idx+1, column=6, value=record.get('total_late_count', 0)).border = thin_border
            ws.cell(row=idx+1, column=7, value=record.get('late_within_10_count', 0)).border = thin_border
            ws.cell(row=idx+1, column=8, value=record.get('late_over_10_count', 0)).border = thin_border
            ws.cell(row=idx+1, column=9, value=record.get('late_over_60_count', 0)).border = thin_border
            ws.cell(row=idx+1, column=10, value=record.get('morning_missing_count', 0)).border = thin_border
            ws.cell(row=idx+1, column=11, value=record.get('evening_missing_count', 0)).border = thin_border
            ws.cell(row=idx+1, column=12, value=record.get('early_leave_count', 0)).border = thin_border
            ws.cell(row=idx+1, column=13, value=float(record.get('total_deduction', 0))).border = thin_border
            ws.cell(row=idx+1, column=14, value=record.get('attendance_month', '')).border = thin_border
        
        # 列宽
        col_widths = [6, 10, 18, 10, 10, 8, 8, 8, 8, 8, 8, 8, 10, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 合计行
        total_row = len(records) + 2
        ws.cell(row=total_row, column=1, value="合计")
        ws.cell(row=total_row, column=6, value=sum(r.get('total_late_count', 0) for r in records))
        ws.cell(row=total_row, column=13, value=sum(float(r.get('total_deduction', 0)) for r in records))
        
        for col in range(1, 15):
            ws.cell(row=total_row, column=col).font = Font(bold=True)
            ws.cell(row=total_row, column=col).border = thin_border
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        month_str = month.replace('-', '') if month else 'all'
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"attendance_deduction_{month_str}_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"[Attendance] 导出失败: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
