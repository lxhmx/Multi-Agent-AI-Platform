"""
财务加班管理 API
提供员工加班数据的上传、查询、管理功能
"""
import sys
import io
from datetime import datetime
from pathlib import Path
from typing import Optional, List

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 添加项目根目录到路径
PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from database.mysql_client import MySQLClient
from api.financial.overtime_calculator import process_attendance

# 创建路由器
router = APIRouter(prefix="/api/financial", tags=["财务管理"])


# ==================== 数据模型 ====================

class OvertimeRecord(BaseModel):
    """加班记录模型"""
    employee_name: str
    job_title: str
    job_level: str
    overtime_hours: int
    overtime_minutes: int
    overtime_days: int
    overtime_rate: float
    overtime_amount: float
    overtime_detail: str
    attendance_month: str


class OvertimeQueryParams(BaseModel):
    """查询参数"""
    page: int = 1
    page_size: int = 20
    month: Optional[str] = None
    keyword: Optional[str] = None


# ==================== 工具函数 ====================

def extract_month_from_filename(filename: str) -> Optional[str]:
    """从文件名中提取月份信息"""
    import re
    
    patterns = [
        r'(\d{4})年(\d{1,2})月',
        r'(\d{4})-(\d{1,2})',
        r'(\d{4})(\d{2})',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            year, month = match.groups()
            return f"{year}-{int(month):02d}"
    
    month_match = re.search(r'(\d{1,2})月', filename)
    if month_match:
        month = int(month_match.group(1))
        year = datetime.now().year
        return f"{year}-{month:02d}"
    
    return None


def save_overtime_records(records: List[dict], attendance_month: str) -> dict:
    """保存加班记录到数据库，当月数据已存在则覆盖更新"""
    if not records:
        return {"inserted": 0, "updated": 0}
    
    inserted = 0
    updated = 0
    
    upsert_sql = """
    INSERT INTO employee_overtime 
        (employee_name, job_title, job_level, overtime_hours, overtime_minutes, 
         overtime_days, overtime_rate, overtime_amount, overtime_detail, attendance_month, row_order)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ON DUPLICATE KEY UPDATE
        job_title = VALUES(job_title),
        job_level = VALUES(job_level),
        overtime_hours = VALUES(overtime_hours),
        overtime_minutes = VALUES(overtime_minutes),
        overtime_days = VALUES(overtime_days),
        overtime_rate = VALUES(overtime_rate),
        overtime_amount = VALUES(overtime_amount),
        overtime_detail = VALUES(overtime_detail),
        row_order = VALUES(row_order),
        updated_at = CURRENT_TIMESTAMP
    """
    
    try:
        with MySQLClient() as db:
            for record in records:
                params = (
                    record['姓名'],
                    record['职务'],
                    record['职级'],
                    record['加班时长(小时)'],
                    record['加班时长(分钟)'],
                    record['加班天数'],
                    record['加班费用标准'],
                    record['加班总金额'],
                    record['详情'],
                    attendance_month,
                    record.get('行序号', 0)
                )
                affected = db.execute_update(upsert_sql, params)
                if affected == 1:
                    inserted += 1
                else:
                    updated += 1
        
        return {"inserted": inserted, "updated": updated}
    except Exception as e:
        raise Exception(f"保存数据失败: {str(e)}")


# ==================== API 接口 ====================

@router.post("/upload-attendance")
async def upload_attendance(
    file: UploadFile = File(...),
    month: Optional[str] = Form(None)
):
    """上传考勤表并计算加班数据"""
    try:
        if not file.filename.endswith(('.xls', '.xlsx')):
            raise HTTPException(status_code=400, detail="请上传Excel文件(.xls或.xlsx)")
        
        attendance_month = month
        if not attendance_month:
            attendance_month = extract_month_from_filename(file.filename)
        if not attendance_month:
            attendance_month = datetime.now().strftime('%Y-%m')
        
        temp_dir = PROJECT_ROOT / 'temp'
        temp_dir.mkdir(exist_ok=True)
        temp_path = temp_dir / f"attendance_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
        
        content = await file.read()
        with open(temp_path, 'wb') as f:
            f.write(content)
        
        try:
            records, excel_month = process_attendance(str(temp_path))
            
            if not records:
                raise HTTPException(status_code=400, detail="未能从文件中解析出有效数据")
            
            # 优先使用Excel中提取的月份，其次使用用户指定的月份，最后使用文件名或当前月份
            final_month = excel_month or attendance_month
            if not final_month:
                final_month = datetime.now().strftime('%Y-%m')
            
            result = save_overtime_records(records, final_month)
            
            total_hours = sum(r['加班时长(小时)'] for r in records)
            total_amount = sum(r['加班总金额'] for r in records)
            overtime_count = len([r for r in records if r['加班时长(小时)'] > 0])
            
            return {
                "success": True,
                "message": "考勤数据处理完成",
                "data": {
                    "month": final_month,
                    "total_employees": len(records),
                    "overtime_employees": overtime_count,
                    "total_hours": total_hours,
                    "total_amount": total_amount,
                    "inserted": result["inserted"],
                    "updated": result["updated"]
                }
            }
        finally:
            if temp_path.exists():
                temp_path.unlink()
                
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"[Financial] 上传处理失败: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")


@router.get("/overtime-records")
async def get_overtime_records(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    month: Optional[str] = Query(None, description="考勤月份(YYYY-MM)"),
    keyword: Optional[str] = Query(None, description="搜索关键词(姓名/职务)"),
    sort_field: Optional[str] = Query(None, description="排序字段"),
    sort_order: Optional[str] = Query(None, description="排序方向(asc/desc)")
):
    """获取加班记录列表"""
    try:
        conditions = []
        params = []
        
        if month:
            conditions.append("attendance_month = %s")
            params.append(month)
        
        if keyword:
            conditions.append("(employee_name LIKE %s OR job_title LIKE %s)")
            params.extend([f"%{keyword}%", f"%{keyword}%"])
        
        where_clause = " AND ".join(conditions) if conditions else "1=1"
        
        # 构建排序子句
        order_clause = "attendance_month DESC, row_order ASC"
        if sort_field and sort_order:
            # 验证排序字段，防止SQL注入
            allowed_fields = ['overtime_hours', 'overtime_days', 'overtime_amount', 'employee_name']
            if sort_field in allowed_fields:
                order_direction = 'ASC' if sort_order.lower() == 'asc' else 'DESC'
                order_clause = f"{sort_field} {order_direction}, row_order ASC"
        
        count_sql = f"SELECT COUNT(*) as total FROM employee_overtime WHERE {where_clause}"
        
        offset = (page - 1) * page_size
        data_sql = f"""
            SELECT * FROM employee_overtime 
            WHERE {where_clause}
            ORDER BY {order_clause}
            LIMIT %s OFFSET %s
        """
        
        with MySQLClient() as db:
            count_result = db.execute_query(count_sql, tuple(params))
            total = count_result[0]['total'] if count_result else 0
            
            data_params = tuple(params) + (page_size, offset)
            records = db.execute_query(data_sql, data_params)
        
        return {
            "success": True,
            "data": records,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": (total + page_size - 1) // page_size
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


@router.get("/overtime-stats")
async def get_overtime_stats(
    month: Optional[str] = Query(None, description="考勤月份(YYYY-MM)")
):
    """获取加班统计数据"""
    try:
        with MySQLClient() as db:
            if month:
                stats_sql = """
                    SELECT 
                        COUNT(*) as total_employees,
                        SUM(CASE WHEN overtime_hours > 0 THEN 1 ELSE 0 END) as overtime_employees,
                        SUM(overtime_hours) as total_hours,
                        SUM(overtime_days) as total_days,
                        SUM(overtime_amount) as total_amount,
                        AVG(overtime_hours) as avg_hours
                    FROM employee_overtime
                    WHERE attendance_month = %s
                """
                stats = db.execute_query(stats_sql, (month,))
            else:
                stats_sql = """
                    SELECT 
                        COUNT(*) as total_employees,
                        SUM(CASE WHEN overtime_hours > 0 THEN 1 ELSE 0 END) as overtime_employees,
                        SUM(overtime_hours) as total_hours,
                        SUM(overtime_days) as total_days,
                        SUM(overtime_amount) as total_amount,
                        AVG(overtime_hours) as avg_hours
                    FROM employee_overtime
                """
                stats = db.execute_query(stats_sql)
            
            months_sql = """
                SELECT DISTINCT attendance_month 
                FROM employee_overtime 
                ORDER BY attendance_month DESC
            """
            months = db.execute_query(months_sql)
            
            # 将具体职级归类到大类：P、M、D、实习
            level_stats_sql = """
                SELECT 
                    CASE 
                        WHEN UPPER(job_level) LIKE 'P%%' AND job_level NOT LIKE '%%实习%%' THEN 'P'
                        WHEN UPPER(job_level) LIKE 'M%%' THEN 'M'
                        WHEN UPPER(job_level) LIKE 'D%%' THEN 'D'
                        WHEN job_level LIKE '%%实习%%' THEN '实习'
                        ELSE '其他'
                    END as level_category,
                    COUNT(*) as count,
                    SUM(overtime_hours) as total_hours,
                    SUM(overtime_amount) as total_amount
                FROM employee_overtime
                {}
                GROUP BY level_category
                ORDER BY FIELD(level_category, 'P', 'M', 'D', '实习', '其他')
            """.format("WHERE attendance_month = %s" if month else "")
            
            level_stats_raw = db.execute_query(level_stats_sql, (month,) if month else None)
            # 转换字段名以兼容前端
            level_stats = [{'job_level': r['level_category'], 'count': r['count'], 
                           'total_hours': r['total_hours'], 'total_amount': r['total_amount']} 
                          for r in level_stats_raw] if level_stats_raw else []
            
            top_sql = """
                SELECT employee_name, job_title, job_level, overtime_hours, overtime_amount
                FROM employee_overtime
                {}
                ORDER BY overtime_hours DESC
                LIMIT 10
            """.format("WHERE attendance_month = %s" if month else "")
            
            top_employees = db.execute_query(top_sql, (month,) if month else None)
        
        return {
            "success": True,
            "data": {
                "stats": stats[0] if stats else {},
                "months": [m['attendance_month'] for m in months],
                "level_stats": level_stats,
                "top_employees": top_employees
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"统计失败: {str(e)}")


@router.delete("/overtime-records")
async def delete_overtime_records(
    ids: Optional[List[int]] = Query(None, description="要删除的记录ID列表"),
    month: Optional[str] = Query(None, description="要删除的月份数据")
):
    """删除加班记录"""
    try:
        if not ids and not month:
            raise HTTPException(status_code=400, detail="请提供要删除的记录ID或月份")
        
        with MySQLClient() as db:
            if ids:
                placeholders = ','.join(['%s'] * len(ids))
                sql = f"DELETE FROM employee_overtime WHERE id IN ({placeholders})"
                affected = db.execute_update(sql, tuple(ids))
            else:
                sql = "DELETE FROM employee_overtime WHERE attendance_month = %s"
                affected = db.execute_update(sql, (month,))
        
        return {
            "success": True,
            "message": f"成功删除 {affected} 条记录"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


@router.get("/overtime-records/export")
async def export_overtime_records(
    month: Optional[str] = Query(None, description="考勤月份(YYYY-MM)"),
    keyword: Optional[str] = Query(None, description="搜索关键词(姓名/职务)"),
    sort_field: Optional[str] = Query(None, description="排序字段"),
    sort_order: Optional[str] = Query(None, description="排序方向(asc/desc)")
):
    """导出加班记录为Excel"""
    try:
        conditions = []
        params = []
        
        if month:
            conditions.append("attendance_month = %s")
            params.append(month)
        
        if keyword:
            conditions.append("(employee_name LIKE %s OR job_title LIKE %s)")
            params.extend([f"%{keyword}%", f"%{keyword}%"])
        
        where_clause = " AND ".join(conditions) if conditions else "1=1"
        
        # 构建排序子句
        order_clause = "attendance_month DESC, row_order ASC"
        if sort_field and sort_order:
            allowed_fields = ['overtime_hours', 'overtime_days', 'overtime_amount', 'employee_name']
            if sort_field in allowed_fields:
                order_direction = 'ASC' if sort_order.lower() == 'asc' else 'DESC'
                order_clause = f"{sort_field} {order_direction}, row_order ASC"
        
        data_sql = f"""
            SELECT * FROM employee_overtime 
            WHERE {where_clause}
            ORDER BY {order_clause}
        """
        
        with MySQLClient() as db:
            records = db.execute_query(data_sql, tuple(params))
        
        if not records:
            raise HTTPException(status_code=404, detail="没有数据可导出")
        
        # 创建Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "加班记录"
        
        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 表头
        headers = ['序号', '姓名', '职务', '职级', '加班时长(小时)', '加班天数', '费用标准(元/小时)', '加班金额', '考勤月份', '加班明细']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 数据行
        for idx, record in enumerate(records, 1):
            ws.cell(row=idx+1, column=1, value=idx).border = thin_border
            ws.cell(row=idx+1, column=2, value=record.get('employee_name', '')).border = thin_border
            ws.cell(row=idx+1, column=3, value=record.get('job_title', '')).border = thin_border
            ws.cell(row=idx+1, column=4, value=record.get('job_level', '')).border = thin_border
            ws.cell(row=idx+1, column=5, value=record.get('overtime_hours', 0)).border = thin_border
            ws.cell(row=idx+1, column=6, value=record.get('overtime_days', 0)).border = thin_border
            ws.cell(row=idx+1, column=7, value=record.get('overtime_rate', 0)).border = thin_border
            ws.cell(row=idx+1, column=8, value=record.get('overtime_amount', 0)).border = thin_border
            ws.cell(row=idx+1, column=9, value=record.get('attendance_month', '')).border = thin_border
            ws.cell(row=idx+1, column=10, value=record.get('overtime_detail', '')).border = thin_border
        
        # 列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 60
        
        # 合计行
        total_row = len(records) + 2
        ws.cell(row=total_row, column=1, value="合计")
        ws.cell(row=total_row, column=5, value=sum(r.get('overtime_hours', 0) for r in records))
        ws.cell(row=total_row, column=6, value=sum(r.get('overtime_days', 0) for r in records))
        ws.cell(row=total_row, column=8, value=sum(r.get('overtime_amount', 0) for r in records))
        
        for col in range(1, 11):
            ws.cell(row=total_row, column=col).font = Font(bold=True)
            ws.cell(row=total_row, column=col).border = thin_border
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名（纯英文，避免编码问题）
        month_str = month.replace('-', '') if month else 'all'
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"overtime_records_{month_str}_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"[Financial] 导出失败: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
