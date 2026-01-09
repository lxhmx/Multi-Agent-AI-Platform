"""
员工加班时长计算脚本
用法: python api/financial/overtime_calculator.py <考勤表路径> [输出文件路径]
示例: python api/financial/overtime_calculator.py test_downloads/1111.xlsx output/加班统计.xlsx

规则:
1. 周末(星期六、星期日)不算加班
2. 每天加班必须满1小时才计入，不满1小时不算
3. 超过1小时的部分按整小时计算（向下取整）
   例如: 加班1小时40分 -> 计1小时

职级规则:
- P开头: 固定工时，09:00-18:00，加班=超过18:00下班的时间
- M开头: 弹性工时，一天需上满9小时，加班=工作时长超过9小时的部分
- D开头: 弹性工时，一天需上满8.5小时，加班=工作时长超过8.5小时的部分
- 实习生: 按P职级规则计算（固定工时）
"""

import openpyxl
import re
import sys
import os
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ============ 配置参数 ============
MIN_OVERTIME_MINUTES = 60       # 最少加班分钟数（不满此数不计入）
DATA_START_ROW = 3              # 数据起始行（跳过表头）
HEADER_ROW = 1                  # 表头行
DATE_INFO_ROW = 2               # 日期信息行（包含星期几）

# 列名匹配关键字（用于动态查找列位置）
COLUMN_KEYWORDS = {
    'name': ['姓名', 'name'],
    'job': ['职务', 'job', '岗位'],
    'level': ['职级', 'level', '对应职级', '对应职 级'],
    'overtime_rate': ['加班费用标准', '费用标准', '加班标准'],
    'month_info': ['考勤', '月份']
}

# 职级对应的工时规则
LEVEL_RULES = {
    'P': {'type': 'fixed', 'end_time': '18:00', 'default_start': '09:00'},
    'M': {'type': 'flexible', 'work_hours': 9, 'default_start': '09:30'},
    'D': {'type': 'flexible', 'work_hours': 8.5, 'default_start': '09:30'},
    '实习': {'type': 'fixed', 'end_time': '18:00', 'default_start': '09:00'},
    'default': {'type': 'fixed', 'end_time': '18:00', 'default_start': '09:00'},
}


def get_level_rule(level: str) -> dict:
    """根据职级获取对应的工时规则"""
    if not level:
        return LEVEL_RULES['default']
    
    level_upper = level.upper()
    if level_upper.startswith('P') or level_upper.startswith('实习'):
        return LEVEL_RULES['P']
    elif level_upper.startswith('M'):
        return LEVEL_RULES['M']
    elif level_upper.startswith('D'):
        return LEVEL_RULES['D']
    elif '实习' in level:
        return LEVEL_RULES['实习']
    return LEVEL_RULES['default']


def extract_month_from_cell(cell_value: str) -> str:
    """
    从单元格内容中提取月份信息
    支持格式: "2025年11月考勤", "2024年1月", "2025-11考勤" 等
    返回: YYYY-MM 格式的月份字符串，如 "2025-11"
    """
    if not cell_value:
        return None
    
    cell_str = str(cell_value)
    
    # 匹配 YYYY年MM月 格式
    match = re.search(r'(\d{4})年(\d{1,2})月', cell_str)
    if match:
        year, month = match.groups()
        return f"{year}-{int(month):02d}"
    
    # 匹配 YYYY-MM 格式
    match = re.search(r'(\d{4})-(\d{1,2})', cell_str)
    if match:
        year, month = match.groups()
        return f"{year}-{int(month):02d}"
    
    # 匹配 YYYYMM 格式
    match = re.search(r'(\d{4})(\d{2})', cell_str)
    if match:
        year, month = match.groups()
        return f"{year}-{month}"
    
    return None


def find_column_by_keywords(sheet, row: int, keywords: list) -> int:
    """
    根据关键字列表在指定行查找列索引
    返回: 列索引（1-based），未找到返回None
    """
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=row, column=col).value
        if cell_value:
            cell_str = str(cell_value).strip()
            for keyword in keywords:
                if keyword in cell_str:
                    return col
    return None


def find_date_columns(sheet, date_info_row: int) -> tuple:
    """
    查找日期列的起始和结束位置
    返回: (起始列, 结束列)
    """
    start_col = None
    end_col = None
    
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=date_info_row, column=col).value
        if cell_value:
            cell_str = str(cell_value)
            # 检查是否包含日期信息（数字+星期）
            if re.search(r'\d+', cell_str) and '星期' in cell_str:
                if start_col is None:
                    start_col = col
                end_col = col
    
    return start_col, end_col


def get_column_mapping(sheet) -> dict:
    """
    动态获取列映射关系
    返回: 包含各列索引的字典
    """
    mapping = {
        'name_col': find_column_by_keywords(sheet, HEADER_ROW, COLUMN_KEYWORDS['name']),
        'job_col': find_column_by_keywords(sheet, HEADER_ROW, COLUMN_KEYWORDS['job']),
        'level_col': find_column_by_keywords(sheet, HEADER_ROW, COLUMN_KEYWORDS['level']),
        'overtime_rate_col': find_column_by_keywords(sheet, HEADER_ROW, COLUMN_KEYWORDS['overtime_rate']),
    }
    
    # 查找日期列范围
    date_start, date_end = find_date_columns(sheet, DATE_INFO_ROW)
    mapping['date_start_col'] = date_start
    mapping['date_end_col'] = date_end
    
    # 查找月份信息列
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=HEADER_ROW, column=col).value
        if cell_value:
            cell_str = str(cell_value)
            if any(keyword in cell_str for keyword in COLUMN_KEYWORDS['month_info']):
                mapping['month_info_col'] = col
                break
    
    # 验证必需列是否找到
    required_cols = ['name_col', 'level_col', 'date_start_col', 'date_end_col']
    missing_cols = [col for col in required_cols if not mapping.get(col)]
    
    if missing_cols:
        raise ValueError(f"无法找到必需的列: {missing_cols}")
    
    return mapping



def parse_work_time(record: str, level_rule: dict = None) -> tuple:
    """解析打卡记录，提取上下班时间"""
    if not record:
        return None, None
    
    time_matches = re.findall(r'\d{2}:\d{2}', record)
    
    if len(time_matches) >= 2:
        try:
            times = [datetime.strptime(t, '%H:%M') for t in time_matches]
            times.sort()
            return times[0], times[-1]
        except ValueError:
            pass
    elif len(time_matches) == 1 and level_rule and 'default_start' in level_rule:
        try:
            default_start = datetime.strptime(level_rule['default_start'], '%H:%M')
            end_time = datetime.strptime(time_matches[0], '%H:%M')
            if end_time > default_start:
                return default_start, end_time
        except ValueError:
            pass
    
    return None, None


def parse_explicit_overtime(record: str) -> int:
    """解析显式标注的加班时长，返回分钟数"""
    if not record:
        return 0
    
    overtime_match = re.search(r'加班(\d+\.?\d*)小时', record)
    if overtime_match:
        return int(float(overtime_match.group(1)) * 60)
    
    overtime_day_match = re.search(r'加班(\d+)天', record)
    if overtime_day_match:
        return int(overtime_day_match.group(1)) * 8 * 60
    
    return 0


def is_skip_record(record: str) -> bool:
    """判断是否跳过该记录"""
    if not record:
        return True
    skip_keywords = ['休息', '请假', '--', '事假', '病假', '年休假', '产检假']
    return any(keyword in record for keyword in skip_keywords)


def is_weekend(day_info: str) -> bool:
    """判断是否为周末"""
    if not day_info:
        return False
    return '星期六' in str(day_info) or '星期日' in str(day_info)


def calculate_daily_overtime(start_time, end_time, level_rule: dict) -> int:
    """根据职级规则计算单日加班时长（分钟）"""
    if not start_time or not end_time:
        return 0
    
    overtime_minutes = 0
    
    if level_rule['type'] == 'fixed':
        standard_end = datetime.strptime(level_rule['end_time'], '%H:%M')
        if end_time > standard_end:
            overtime_minutes = (end_time - standard_end).seconds // 60
    else:
        work_minutes = (end_time - start_time).seconds // 60
        standard_minutes = int(level_rule['work_hours'] * 60)
        if work_minutes > standard_minutes:
            overtime_minutes = work_minutes - standard_minutes
    
    if overtime_minutes < MIN_OVERTIME_MINUTES:
        return 0
    
    return (overtime_minutes // 60) * 60


def process_attendance(file_path: str) -> tuple:
    """
    处理考勤表，计算每个员工的加班时长和加班费用
    返回: (员工加班数据列表, 考勤月份)
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    # 动态获取列映射
    try:
        col_map = get_column_mapping(sheet)
        print(f"列映射: 姓名={col_map['name_col']}, 职务={col_map.get('job_col')}, "
              f"职级={col_map['level_col']}, 加班标准={col_map.get('overtime_rate_col')}, "
              f"日期列={col_map['date_start_col']}-{col_map['date_end_col']}")
    except ValueError as e:
        print(f"错误: {e}")
        raise
    
    # 从Excel中提取考勤月份
    attendance_month = None
    if col_map.get('month_info_col'):
        month_cell_value = sheet.cell(row=HEADER_ROW, column=col_map['month_info_col']).value
        attendance_month = extract_month_from_cell(month_cell_value)
    
    results = []
    
    for row_idx in range(DATA_START_ROW, sheet.max_row + 1):
        name = sheet.cell(row=row_idx, column=col_map['name_col']).value
        job = sheet.cell(row=row_idx, column=col_map.get('job_col')).value if col_map.get('job_col') else None
        level = sheet.cell(row=row_idx, column=col_map['level_col']).value
        overtime_rate = sheet.cell(row=row_idx, column=col_map.get('overtime_rate_col')).value if col_map.get('overtime_rate_col') else None
        
        if not name:
            continue
        
        try:
            overtime_rate = float(overtime_rate) if overtime_rate else 0
        except (ValueError, TypeError):
            overtime_rate = 0
        
        level_rule = get_level_rule(level)
        total_overtime_minutes = 0
        overtime_days = 0
        daily_details = []
        
        for col_idx in range(col_map['date_start_col'], col_map['date_end_col'] + 1):
            day_info = sheet.cell(row=DATE_INFO_ROW, column=col_idx).value
            
            if is_weekend(day_info):
                continue
            
            record = sheet.cell(row=row_idx, column=col_idx).value
            if not record:
                continue
            
            record = str(record)
            if is_skip_record(record):
                continue
            
            daily_overtime = 0
            explicit_overtime = parse_explicit_overtime(record)
            if explicit_overtime > 0:
                if explicit_overtime >= MIN_OVERTIME_MINUTES:
                    daily_overtime = (explicit_overtime // 60) * 60
            else:
                start_time, end_time = parse_work_time(record, level_rule)
                if start_time and end_time:
                    daily_overtime = calculate_daily_overtime(start_time, end_time, level_rule)
            
            if daily_overtime > 0:
                total_overtime_minutes += daily_overtime
                overtime_days += 1
                day_num = str(day_info).split('\n')[0] if day_info else ''
                daily_details.append(f"{day_num}日:{daily_overtime//60}h")
        
        overtime_hours = total_overtime_minutes // 60
        overtime_amount = overtime_hours * overtime_rate
        
        results.append({
            '姓名': name,
            '职务': job or '',
            '职级': level or '',
            '加班时长(小时)': overtime_hours,
            '加班时长(分钟)': total_overtime_minutes,
            '加班天数': overtime_days,
            '加班费用标准': overtime_rate,
            '加班总金额': overtime_amount,
            '详情': ', '.join(daily_details) if daily_details else '',
            '行序号': row_idx  # 添加原始行号
        })
    
    return results, attendance_month


def export_to_excel(results: list, output_path: str):
    """导出结果到Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "加班统计"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ['序号', '姓名', '职务', '职级', '加班时长(小时)', '加班总金额', '加班明细']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for idx, row_data in enumerate(results, 1):
        ws.cell(row=idx+1, column=1, value=idx).border = thin_border
        ws.cell(row=idx+1, column=2, value=row_data['姓名']).border = thin_border
        ws.cell(row=idx+1, column=3, value=row_data['职务']).border = thin_border
        ws.cell(row=idx+1, column=4, value=row_data['职级']).border = thin_border
        ws.cell(row=idx+1, column=5, value=row_data['加班时长(小时)']).border = thin_border
        ws.cell(row=idx+1, column=6, value=row_data['加班总金额']).border = thin_border
        ws.cell(row=idx+1, column=7, value=row_data['详情']).border = thin_border
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 80
    
    total_row = len(results) + 2
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=5, value=sum(r['加班时长(小时)'] for r in results))
    ws.cell(row=total_row, column=6, value=sum(r['加班总金额'] for r in results))
    
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).border = thin_border
    
    wb.save(output_path)
    print(f"结果已导出到: {output_path}")


def print_summary(results: list):
    """打印统计摘要"""
    overtime_employees = [r for r in results if r['加班时长(小时)'] > 0]
    
    print("\n" + "=" * 70)
    print("加班统计摘要")
    print("=" * 70)
    print(f"总员工数: {len(results)}")
    print(f"有加班记录的员工: {len(overtime_employees)}")
    print(f"总加班时长: {sum(r['加班时长(小时)'] for r in results)} 小时")
    print(f"总加班天数: {sum(r['加班天数'] for r in results)} 天")
    
    print("\n按职级统计:")
    print("-" * 50)
    level_stats = {}
    for r in results:
        level = r['职级'] or '未知'
        if level not in level_stats:
            level_stats[level] = {'count': 0, 'hours': 0}
        level_stats[level]['count'] += 1
        level_stats[level]['hours'] += r['加班时长(小时)']
    
    for level in sorted(level_stats.keys()):
        stats = level_stats[level]
        print(f"  {level:<12} {stats['count']:>3}人  共{stats['hours']:>4}小时")
    
    print("\n加班时长 TOP 10:")
    print("-" * 60)
    sorted_results = sorted(overtime_employees, key=lambda x: x['加班时长(小时)'], reverse=True)[:10]
    for i, r in enumerate(sorted_results, 1):
        print(f"{i:2}. {r['姓名']:<8} {r['职务']:<12} {r['职级']:<8} {r['加班时长(小时)']:>3}小时 ({r['加班天数']}天)")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("错误: 请提供考勤表文件路径")
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    if not os.path.exists(input_file):
        print(f"错误: 文件不存在 - {input_file}")
        sys.exit(1)
    
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        output_dir = os.path.dirname(input_file) or '.'
        output_file = os.path.join(output_dir, f"{base_name}_加班统计.xlsx")
    
    print(f"正在处理: {input_file}")
    results, attendance_month = process_attendance(input_file)
    print(f"考勤月份: {attendance_month or '未识别'}")
    print_summary(results)
    export_to_excel(results, output_file)


if __name__ == "__main__":
    main()
